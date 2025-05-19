"""
Created on 18/05/2025
@author: frposada

Funciones para manejo de archivos y datos, incluyendo protección de archivos Excel.
"""
#%%-------------------------------- Librerias ----------------------------------
import os                                                                       # proporciona funciones para interactuar con el sistema operativo (biblioteca estandar)
import logging                                                                  # Permite generar un sistema de registro de eventos (biblioteca estandar)
import pandas as pd                                                             # para el manejo y analisis de estructuras de datos
from openpyxl import load_workbook                                              # para leer y escribir archivos de Excel
from openpyxl.styles import Protection                                          # para aplicar estilos a las celdas de Excel

#%%---------------------------- funciones y clases -----------------------------
def create_file(path: str, content: str):
  """Crea un archivo y escribe contenido en él.

  Args:
    - path (str): Ruta completa del archivo (incluyendo nombre y extensión).
    - content (str): Contenido a escribir en el archivo.
  Example:
    create_file("C:/Users/usuario/archivo.txt", "Contenido del archivo")
  """
  logging.info(f"Creando archivo: {path}")                                      # mensaje de control
  with open(path, "w") as file:                                                 # Abre el archivo en modo escritura
    file.write(str(content))                                                    # Escribe el contenido en el archivo

def remove_files(path: str):
  """Elimina todos los archivos dentro de la carpeta especificada.

  Args:
    - path (str): Ruta de la carpeta cuyos archivos se eliminarán.
  Example:
    remove_files("C:/Users/usuario/Carpeta")
  """
  logging.info(f"Eliminando archivos en la carpeta: {path}")                    # mensaje de control
  for f in os.listdir(path):                                                    # Itera sobre todos los archivos en la carpeta
    os.remove(os.path.join(path, f))                                            # Elimina cada archivo

def read_files(path: str, extensions: list=None, files: list=None, separator: str="|", header: str="infer", consolidate: bool=False):
    """Lee archivos de una carpeta según los criterios especificados y opcionalmente los consolida.

    Args:
      - path (str): Ruta de la carpeta con los archivos.
      - extensions (list, optional): Lista de extensiones de archivos a leer (por ejemplo, ["xlsx", "csv", "txt"]). Si es None, lee todos los archivos.
      - files (list, optional): Lista de nombres de archivos específicos a leer. Si es None, lee todos los archivos válidos según las extensiones.
      - separator (str, optional): Separador para archivos de texto (default: "|").
      - header (str, optional): Fila de encabezado para archivos CSV (default: "infer"). infer -> infiere el encabezado (primera fila), None -> no hay encabezado.
      - consolidate (bool, optional): Consolida los archivos en un único DataFrame (default: False).

    Returns:
      - list | pandas.DataFrame: Lista de DataFrames o un DataFrame consolidado.

    Example:
      - Leer todos los archivos de la carpeta
      dataframes = read_files("C:/Users/usuario/Carpeta")

      - Leer solo archivos con extensiones específicas
      dataframes = read_files("C:/Users/usuario/Carpeta", extensions=["xlsx", "csv"])

      - Leer archivos específicos
      dataframes = read_files("C:/Users/usuario/Carpeta", files=["archivo1.csv", "archivo2.xlsx"])

      - Consolidar todos los archivos en un único DataFrame
      consolidated_df = read_files("C:/Users/usuario/Carpeta", consolidate=True)
    """
    logging.info(f"Leyendo archivos en la carpeta: {path}")                     # mensaje de control
    all_files = os.listdir(path)                                                # Obtener todos los archivos de la carpeta

    if extensions:                                                              # si se especifica una lista de extensiones
      extensions = [ext.lower() for ext in extensions]                          # Asegurar que las extensiones sean minúsculas
      all_files = [f for f in all_files if any(f.lower().endswith(f".{ext}") for ext in extensions)] # Filtrar por extensiones

    if files:                                                                   # si se especifica una lista de archivos
      all_files = [f for f in all_files if f in files]                          # Filtrar por archivos específicos

    dataframes = []                                                             # Lista para almacenar DataFrames
    for file in all_files:                                                      # Iterar sobre los archivos filtrados
      file_path = os.path.join(path, file)                                      # Obtener la ruta completa del archivo
      try:
        if file.lower().endswith(".xlsx"):                                      # Si el archivo es de Excel
          dataframes.append(pd.read_excel(file_path))                           # Leer el archivo de Excel
        elif file.lower().endswith((".csv", ".txt")):                           # Si el archivo es CSV o TXT
          dataframes.append(pd.read_csv(file_path, sep=separator, header=header)) # Leer el archivo CSV o TXT
        else:
          logging.warning(f"Formato no soportado para el archivo: {file}")      # mensaje de advertencia si el formato no es soportado
          continue                                                              # Continuar con la siguiente iteración
        logging.info(f"Archivo {file} leído correctamente.")                    # mensaje de control
      except Exception as e:                                                    # Manejo de excepciones
        logging.warning(f"No se pudo leer el archivo {file}: {e}")              # mensaje de advertencia si no se pudo leer el archivo

    return pd.concat(dataframes, ignore_index=True) if consolidate else dataframes # Retornar un DataFrame consolidado o una lista de DataFrames

def searchv(df: pd.DataFrame, search_field: str, search_value: any, return_field: str=None):
    """Busca un valor en un DataFrame y retorna los resultados.

    Args:
      - df (pandas.DataFrame): DataFrame donde buscar.
      - search_field (str): Columna donde buscar el valor.
      - search_value (any): Valor a buscar.
      - return_field (str): Columna a retornar (default: None).
        1) return_field -> valor: default -> traera todos los campos (trae todo el registro)
        2) return_field -> valor: string -> traera el campo especificado
        3) return_field -> valor: lista de strings -> traera todos los campos especificados
    Returns:
      - pandas.DataFrame | pandas.Series: Resultados filtrados.
    Example:
      - Trae todos los campos del registro con ID 123
      resultados = searchv(df, "ID", 123)

      - Trae solo el campo "Nombre" del registro con ID 123
      nombres = searchv(df, "ID", 123, "Nombre")

      - Trae los campos "Nombre" y "Edad" del registro con ID 123
      datos = searchv(df, "ID", 123, ["Nombre", "Edad"]) # Trae los campos "Nombre" y "Edad" del registro con ID 123
    """
    logging.info(f"Buscando informacion")                                       # mensaje de control
    if return_field:                                                            # condicion para verificar si el parametro tiene un valor asignado (es decir, no es None)
      result = df[df[search_field] == search_value][return_field]               # se busca el "search_value" en "search_field" y se trae lo que encuentre en el "return_field"
      result = pd.DataFrame(result)                                             # conversion del arreglo a df - para retornar el mismo tipo de dato en ambos casos. Ademas permite organizar las dimensiones
    else:
      result = df[df[search_field] == search_value]                             # se busca el "search_value" en el "search_field" y se trae todo el registro

    if result.empty:                                                            # condicion para verificar si la serie esta vacia (esto sucede si el radicado analizado no tiene el Sig_estados "usuario entrada")
      logging.info("No se encontro " + str(search_value) + " en el campo " + search_field) # mensaje de control
    else:
      result = result.reset_index(drop = True)                                  # reinicio de indice
    return result                                                               # retorno de funcion

def protect_excel(file_path: str, sheet_name: str = "Sheet1", password: str = "frank", unlock_columns: list = None, protect_all_sheets: bool = False):
  """Protege un libro de excel o una hoja en especifico, adicionalmente permite desbloquear columnas especificas.
  nota: Si se protege todo el libro y se indican columnas a desbloquear, estas se desbloquearán en todas las hojas; si alguna columna no existe en una hoja, se mostrará una advertencia.

  Args:
    - file_path (str): Ruta del archivo de Excel a proteger.
    - sheet_name (str): Nombre de la hoja a proteger (default: "Sheet1").
    - password (str): Contraseña para proteger la hoja (default: "frank").
    - unlock_columns (list): Lista de nombres de columnas a desbloquear (default: None).
    - protect_all_sheets (bool): Si True, protege todas las hojas del libro (default: False).
  Example:
    - protege una hoja especifica
    protect_excel("C:/Users/usuario/archivo.xlsx", sheet_name="Hoja1", password="mi_contraseña")

    - protege una hoja especifica y desbloquea columnas especificas
    protect_excel("C:/Users/usuario/archivo.xlsx", sheet_name="Hoja1", password="mi_contraseña", unlock_columns=["Columna1", "Columna2"])

    - protege todas las hojas del libro
    protect_excel("C:/Users/usuario/archivo.xlsx", password="mi_contraseña", protect_all_sheets=True)

    - protege todas las hojas del libro y desbloquea columnas especificas
    protect_excel("C:/Users/usuario/archivo.xlsx", password="mi_contraseña", unlock_columns=["Columna1", "Columna2"], protect_all_sheets=True)
  """
  logging.info(f"Protegiendo archivo: {file_path}")                             # mensaje de control
  try:
    wb = load_workbook(file_path)                                               # cargar el archivo
  except Exception as e:                                                        # Manejo de excepciones
    logging.error(f"No se pudo abrir el archivo: {file_path}. Error: {e}")      # mensaje de error si no se pudo abrir el archivo
    return False                                                                # retorno de funcion

  try:
    sheets_to_protect = wb.worksheets if protect_all_sheets else [wb[sheet_name]] # seleccionar las hojas a proteger
    for ws in sheets_to_protect:                                                # iterar sobre las hojas
      # configuracion de proteccion (True = deshabilitar, False = habilitar)
      ws.protection.sheet = False                                               # habilitar proteccion de hoja
      ws.protection.formatCells = True                                          # deshabilitar formateo de celdas
      ws.protection.formatRows = False                                          # habilitar formateo de filas
      ws.protection.formatColumns = False                                       # habilitar formateo de columnas
      ws.protection.insertColumns = True                                        # deshabilitar insertar columnas
      ws.protection.insertRows = True                                           # deshabilitar insertar filas
      ws.protection.insertHyperlinks = True                                     # deshabilitar insertar hiperenlaces
      ws.protection.deleteColumns = True                                        # deshabilitar eliminar columnas
      ws.protection.deleteRows = True                                           # deshabilitar eliminar filas
      ws.protection.selectLockedCells = False                                   # habilitar seleccion de celdas bloqueadas
      ws.protection.selectUnlockedCells = False                                 # habilitar seleccion de celdas desbloqueadas
      ws.protection.sort = True                                                 # deshabilitar ordenamiento
      ws.protection.autoFilter=False                                            # habilitar autofiltro
      ws.protection.set_password(password)                                      # establecer la contraseña de la hoja

      # Desbloquear columnas específicas
      if unlock_columns:                                                        # si hay columnas a desbloquear
        header = [cell.value for cell in ws[1]]                                 # obtener el encabezado de la hoja
        for col_name in unlock_columns:                                         # iterar sobre las columnas a desbloquear
          if col_name not in header:                                            # si la columna no existe en el encabezado
            logging.warning(f"Columna '{col_name}' no encontrada en la hoja '{ws.title}'") # mensaje de advertencia
            continue                                                            # continuar con la siguiente columna
          col_index = header.index(col_name) + 1                                # obtener el indice de la columna
          for row in ws.iter_rows(min_col=col_index, max_col=col_index):        # iterar sobre las filas de la columna
            for cell in row:                                                    # iterar sobre las celdas de la fila
              cell.protection = Protection(locked=False)                        # desbloquear la celda

    wb.save(file_path)                                                          # guardar el archivo
    logging.info(f"Archivo Excel exportado y protegido: {file_path}")           # mensaje de control
  except Exception as e:                                                        # Manejo de excepciones
    logging.error(f"No se pudo proteger el archivo: {file_path}. Error: {e}")   # mensaje de error